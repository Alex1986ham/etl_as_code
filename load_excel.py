from openpyxl import load_workbook

file_temp = "C:/Users/aldudko/Desktop/Temp/test_import.xlsx"

load_workbook(file_temp, read_only=False, keep_vba=False, data_only=False)

wb = load_workbook(file_temp, read_only=True)

ws = wb.get_sheet_by_name('Tabelle1')

#iter_rows(range_string=None, row_offset=0, column_offset=0)

lists = list()

for row in ws.iter_rows():
    selected_fields = [
                    row[CARD_ID],
                    row[Name],
                    row[IN_Out]
    ]

